#!/usr/bin/env python3
"""Assemble per-task audit results into a single xlsx file.

Reads:
    - The input CSV (for task IDs in original order + task metadata)
    - openclaw_qc_auditor/results/<task_id>.md for each task

Extracts:
    - Qc Score, Selected Error Categories, Auditor Feedback (Rubric Criteria),
      Overall Auditor Feedback (the narrative)

Writes a single xlsx with two sheets:
    - "Audit Results" — one row per task with the 4 audit fields
    - "Summary" — score distribution + tag frequency

Usage:
    python openclaw_qc_auditor/scripts/build_audit_xlsx.py <input_csv> <output_xlsx>
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RESULTS_DIR = Path("openclaw_qc_auditor/results")


def parse_audit(audit_path: Path) -> dict:
    """Extract the 4 audit fields from a results/<task_id>.md file."""
    if not audit_path.exists():
        return {"score": "MISSING", "tags": "", "rubric_feedback": "", "narrative": ""}
    text = audit_path.read_text()

    # Score
    score_m = (
        re.search(r"\*\*Qc Score\*\*\s*\|\s*[`\"]*(\d+|N/?A|Invalidated)[`\"]*", text)
        or re.search(r"Qc Score:\s*[`*]*(\d+|N/?A|Invalidated)", text)
    )
    score = score_m.group(1) if score_m else "?"

    # Selected Error Categories
    tags_m = (
        re.search(r"\*\*Selected Error Categories\*\*\s*\|\s*`(.+?)`\s*\|", text, re.DOTALL)
        or re.search(r"Selected Error Categories:?\s*[`*]*\s*(.+?)(?:\n|$)", text)
    )
    tags = tags_m.group(1).strip() if tags_m else ""
    tags = tags.replace("`", "").strip()
    if tags.lower() in {"(none)", "none", ""}:
        tags = "(no specialization)"

    # Auditor Feedback (Rubric Criteria)
    rfb_m = re.search(
        r"\*\*Auditor Feedback \(Rubric Criteria\)\*\*\s*\|\s*`(.+?)`",
        text,
    )
    rfb = rfb_m.group(1).strip() if rfb_m else "(no specialization)"

    # Overall Auditor Feedback narrative — the block in code fence after "Overall Auditor Feedback"
    nar_m = re.search(
        r"Overall Auditor Feedback\s*\n+```\s*\n(.+?)\n```", text, re.DOTALL
    )
    if not nar_m:
        # Try alternative format: section header then content
        nar_m = re.search(
            r"##\s*Overall Auditor Feedback\s*\n+(.+?)(?=\n##|\Z)",
            text,
            re.DOTALL,
        )
    narrative = nar_m.group(1).strip() if nar_m else ""

    # Confidence block (added v2 — Pass E)
    conf_level = ""
    conf_triggers = ""
    conf_action = ""
    conf_div = ""
    conf_block_m = re.search(r"##\s*Confidence\s*\n(.+?)(?=\n##|\Z)", text, re.DOTALL)
    if conf_block_m:
        block = conf_block_m.group(1)
        lvl_m = re.search(r"Level:\s*(HIGH|MEDIUM|LOW)", block, re.IGNORECASE)
        if lvl_m:
            conf_level = lvl_m.group(1).upper()
        trg_m = re.search(r"Triggers fired:\s*(.+?)(?:\n[A-Z]|\nRecommended|\Z)", block, re.DOTALL)
        if trg_m:
            conf_triggers = trg_m.group(1).strip().replace("\n", " ")
            if conf_triggers.lower() in ("none", "[]", "(none)"):
                conf_triggers = "none"
        act_m = re.search(r"Recommended action:\s*([^\n]+)", block)
        if act_m:
            conf_action = act_m.group(1).strip()
        div_m = re.search(r"Score divergence:\s*(\d+)", block)
        if div_m:
            conf_div = div_m.group(1)

    return {
        "score": score,
        "tags": tags,
        "rubric_feedback": rfb,
        "narrative": narrative,
        "confidence_level": conf_level,
        "confidence_triggers": conf_triggers,
        "confidence_action": conf_action,
        "score_divergence": conf_div,
    }


def autosize_columns(ws, max_width=80):
    """Approximate auto-fit for column widths."""
    for col_cells in ws.columns:
        col = col_cells[0].column_letter
        max_len = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            for line in str(v).splitlines():
                if len(line) > max_len:
                    max_len = len(line)
        ws.column_dimensions[col].width = min(max(10, max_len + 2), max_width)


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("input_csv", help="Path to the input CSV (used for task order + metadata).")
    p.add_argument("output_xlsx", help="Path to write the xlsx file.")
    p.add_argument(
        "--results-dir",
        default=str(RESULTS_DIR),
        help="Directory containing per-task <task_id>.md audits.",
    )
    args = p.parse_args()

    results_dir = Path(args.results_dir)
    df = pd.read_csv(args.input_csv, dtype=str, keep_default_na=False)

    rows = []
    for _, r in df.iterrows():
        tid = r.get("task_id") or r.get("Task Id") or r.get("Task ID")
        attempt = r.get("attempt") or r.get("Attempt") or r.get("Attempt ID", "")
        scen = r.get("scenario_type") or r.get("Scenario Type", "")
        univ = r.get("assigned_universe") or r.get("Assigned Universe", "")
        saf = r.get("has_safety_issues") or r.get("Has Safety Issues", "")
        ft = r.get("failing_tests_count") or r.get("Failing Tests Count", "")
        rc = r.get("final_rubric_criteria_count") or r.get("Final Rubric Criteria Count", "")

        a = parse_audit(results_dir / f"{tid}.md")
        rows.append({
            "task_id": tid,
            "attempt": attempt,
            "universe": univ,
            "scenario_type": scen,
            "has_safety_issues": saf,
            "failing_tests_count": ft,
            "rubric_count": rc,
            "qc_score": a["score"],
            "selected_error_categories": a["tags"],
            "auditor_feedback_rubric_criteria": a["rubric_feedback"],
            "overall_auditor_feedback": a["narrative"],
            "confidence_level": a["confidence_level"],
            "confidence_triggers": a["confidence_triggers"],
            "confidence_action": a["confidence_action"],
            "score_divergence": a["score_divergence"],
        })

    wb = Workbook()

    # === Sheet 1: Audit Results ===
    ws = wb.active
    ws.title = "Audit Results"

    headers = [
        "Task ID",
        "Attempt ID",
        "Universe",
        "Scenario Type",
        "Has Safety Issues",
        "Failing Tests Count",
        "Rubric Criteria Count",
        "QC Score",
        "Confidence",
        "Score Divergence",
        "Review Triggers",
        "Recommended Action",
        "Selected Error Categories",
        "Auditor Feedback (Rubric Criteria)",
        "Overall Auditor Feedback",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    score_color = {
        "5": "C6EFCE",  # green
        "4": "DDEBF7",  # light blue
        "3": "FFEB9C",  # yellow
        "2": "FFC7CE",  # red
        "1": "FFC7CE",
    }
    confidence_color = {
        "HIGH":   "C6EFCE",  # green
        "MEDIUM": "FFEB9C",  # yellow
        "LOW":    "FFC7CE",  # red
    }

    for row in rows:
        ws.append([
            row["task_id"],
            row["attempt"],
            row["universe"],
            row["scenario_type"],
            row["has_safety_issues"],
            row["failing_tests_count"],
            row["rubric_count"],
            row["qc_score"],
            row["confidence_level"],
            row["score_divergence"],
            row["confidence_triggers"],
            row["confidence_action"],
            row["selected_error_categories"],
            row["auditor_feedback_rubric_criteria"],
            row["overall_auditor_feedback"],
        ])
        # Color the score cell (col 8)
        score = str(row["qc_score"])
        if score in score_color:
            cell = ws.cell(row=ws.max_row, column=8)
            cell.fill = PatternFill(start_color=score_color[score], end_color=score_color[score], fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # Color the confidence cell (col 9)
        lvl = (row["confidence_level"] or "").upper()
        if lvl in confidence_color:
            cell = ws.cell(row=ws.max_row, column=9)
            cell.fill = PatternFill(start_color=confidence_color[lvl], end_color=confidence_color[lvl], fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # Wrap text on triggers + tags + narrative
        for col in (11, 13, 15):
            ws.cell(row=ws.max_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    autosize_columns(ws)
    # Freeze header
    ws.freeze_panes = "A2"
    # Wider widths
    ws.column_dimensions["K"].width = 50  # Review Triggers
    ws.column_dimensions["M"].width = 60  # Selected Error Categories
    ws.column_dimensions["O"].width = 80  # Overall Auditor Feedback

    # === Sheet 2: Summary ===
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Total tasks", len(rows)])
    ws2.append([])
    ws2.append(["Score distribution"])
    score_counts = Counter([r["qc_score"] for r in rows])
    ws2.append(["Score", "Count", "% of total"])
    for s in ["5", "4", "3", "2", "1"]:
        n = score_counts.get(s, 0)
        pct = f"{100*n/len(rows):.1f}%" if rows else "0%"
        ws2.append([s, n, pct])
    other = sum(v for k, v in score_counts.items() if k not in {"5", "4", "3", "2", "1"})
    if other:
        ws2.append(["Other/Missing", other, f"{100*other/len(rows):.1f}%"])
    ws2.append([])
    ws2.append(["Tag frequency"])
    ws2.append(["Tag", "Count"])
    tag_counter = Counter()
    for r in rows:
        for m in re.findall(r"\[(?:Fail|Non-Fail|Pass) - [^\]]+\]", r["selected_error_categories"]):
            tag_counter[m] += 1
    for tag, n in sorted(tag_counter.items(), key=lambda x: (-x[1], x[0])):
        ws2.append([tag, n])

    # Style header rows in summary
    for row_idx in (1, 3, 4, 9, 10):
        for cell in ws2[row_idx]:
            if cell.value:
                cell.font = Font(bold=True)
    autosize_columns(ws2)

    # === Sheet 3: Needs Review ===
    # Tasks flagged MEDIUM or LOW confidence — human spot-check queue.
    needs_review = [r for r in rows if (r.get("confidence_level") or "").upper() in ("MEDIUM", "LOW")]
    ws3 = wb.create_sheet("Needs Review")
    ws3.append([
        "Task ID", "Attempt ID", "Universe", "QC Score",
        "Confidence", "Score Divergence", "Review Triggers",
        "Recommended Action", "Selected Error Categories", "Why flagged (narrative head)",
    ])
    for c in ws3[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    # LOW first, then MEDIUM
    needs_review.sort(key=lambda r: (0 if (r.get("confidence_level") or "").upper() == "LOW" else 1, r["task_id"]))
    for r in needs_review:
        ws3.append([
            r["task_id"], r["attempt"], r["universe"], r["qc_score"],
            r["confidence_level"], r["score_divergence"], r["confidence_triggers"],
            r["confidence_action"], r["selected_error_categories"],
            (r["overall_auditor_feedback"] or "")[:500],
        ])
        lvl = (r.get("confidence_level") or "").upper()
        if lvl in confidence_color:
            cell = ws3.cell(row=ws3.max_row, column=5)
            cell.fill = PatternFill(start_color=confidence_color[lvl], end_color=confidence_color[lvl], fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in (7, 9, 10):
            ws3.cell(row=ws3.max_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    ws3.freeze_panes = "A2"
    ws3.column_dimensions["G"].width = 60
    ws3.column_dimensions["I"].width = 50
    ws3.column_dimensions["J"].width = 80
    autosize_columns(ws3, max_width=80)

    # Add Needs Review counts to summary
    low_n = sum(1 for r in rows if (r.get("confidence_level") or "").upper() == "LOW")
    med_n = sum(1 for r in rows if (r.get("confidence_level") or "").upper() == "MEDIUM")
    high_n = sum(1 for r in rows if (r.get("confidence_level") or "").upper() == "HIGH")
    ws2.append([])
    ws2.append(["Confidence distribution"])
    ws2.append(["Level", "Count", "% of total"])
    for lvl, n in (("HIGH", high_n), ("MEDIUM", med_n), ("LOW", low_n)):
        pct = f"{100*n/len(rows):.1f}%" if rows else "0%"
        ws2.append([lvl, n, pct])

    out_path = Path(args.output_xlsx)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    print(f"[ok] Wrote {len(rows)} audits to {out_path}")
    print(f"\nScore distribution:")
    for s in ["5", "4", "3", "2", "1"]:
        n = score_counts.get(s, 0)
        if n:
            print(f"  Score {s}: {n} ({100*n/len(rows):.1f}%)")
    if other:
        print(f"  Other/Missing: {other}")
    print(f"\nTag frequency (top 10):")
    for tag, n in sorted(tag_counter.items(), key=lambda x: (-x[1], x[0]))[:10]:
        print(f"  {n:>2}x {tag}")

    print(f"\nConfidence distribution:")
    print(f"  HIGH:   {high_n} ({100*high_n/len(rows):.1f}%)" if rows else "  HIGH: 0")
    print(f"  MEDIUM: {med_n} ({100*med_n/len(rows):.1f}%)" if rows else "  MEDIUM: 0")
    print(f"  LOW:    {low_n} ({100*low_n/len(rows):.1f}%)" if rows else "  LOW: 0")
    if needs_review:
        print(f"\nNeeds Review queue: {len(needs_review)} tasks (sheet 'Needs Review')")


if __name__ == "__main__":
    main()
