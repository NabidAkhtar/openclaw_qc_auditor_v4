#!/usr/bin/env python3
"""Add a Human-Friendly Feedback column to an audit xlsx.

Reads per-task human feedback from one or more JSON files (each: {task_id: feedback_text})
and merges them into the xlsx as a new rightmost column.

Usage:
    python scripts/add_human_feedback.py \
        <input.xlsx> <output.xlsx> \
        results/_<batch>_audited/_human_feedback_batch_*.json
"""

from __future__ import annotations

import argparse
import glob
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


def load_feedback(json_paths: list[str]) -> dict[str, str]:
    merged: dict[str, str] = {}
    for pattern in json_paths:
        for path in sorted(glob.glob(pattern) or [pattern]):
            p = Path(path)
            if not p.exists():
                print(f"[warn] missing: {p}")
                continue
            with p.open() as f:
                data = json.load(f)
            print(f"[ok]   loaded {p.name}: {len(data)} entries")
            merged.update(data)
    return merged


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("input_xlsx")
    p.add_argument("output_xlsx")
    p.add_argument("feedback_jsons", nargs="+",
                   help="One or more JSON files (or glob patterns) with {task_id: feedback}")
    p.add_argument("--task-id-col", default="Task ID")
    p.add_argument("--column-header", default="Human-Friendly Feedback")
    args = p.parse_args()

    feedback = load_feedback(args.feedback_jsons)
    if not feedback:
        sys.exit("[err] no feedback loaded")
    print(f"[ok]   total: {len(feedback)} feedback entries")

    wb = load_workbook(args.input_xlsx)
    ws = wb["Audit Results"] if "Audit Results" in wb.sheetnames else wb.active

    headers = [c.value for c in ws[1]]
    try:
        tid_col_idx = headers.index(args.task_id_col) + 1
    except ValueError:
        sys.exit(f"[err] task ID column {args.task_id_col!r} not found. Headers: {headers}")

    new_col_idx = ws.max_column + 1
    ws.cell(row=1, column=new_col_idx, value=args.column_header)
    header_cell = ws.cell(row=1, column=new_col_idx)
    header_cell.font = Font(bold=True, color="FFFFFF")
    header_cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    matched = 0
    missing: list[str] = []
    for r in range(2, ws.max_row + 1):
        tid = ws.cell(row=r, column=tid_col_idx).value
        text = feedback.get(str(tid).strip())
        if text is None:
            missing.append(str(tid))
            continue
        cell = ws.cell(row=r, column=new_col_idx, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        matched += 1

    ws.column_dimensions[ws.cell(row=1, column=new_col_idx).column_letter].width = 90

    out = Path(args.output_xlsx)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print(f"\n[done] Wrote {out} with {matched}/{ws.max_row - 1} feedback entries.")
    if missing:
        print(f"[warn] {len(missing)} tasks have no feedback. Sample: {missing[:5]}")


if __name__ == "__main__":
    main()
