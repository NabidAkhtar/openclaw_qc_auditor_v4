#!/usr/bin/env python3
"""Compile a single-tab review xlsx from per-task verdicts.

Produces one sheet (no tabs) with:
  Task ID | Attempt | Universe | Last Attempter Email | QC Score |
  Selected Error Categories | Confidence % | Band | Needs Human Review |
  Why Human Review | Sharp Focus Areas (for human audit) | Brief Audit Summary

Confidence % methodology:
  Start at 100. Subtract weighted penalties for each trigger fired.
  Weights reflect historical miss-rate from the calibration corpus:

    T1  Stage 1<->2 score divergence >= 2        -25%
    T2  Inverse-pair candidates found             -10%
    T3  Named artifacts missing from universe     -15%
    T4  Test infrastructure anomaly               -10%
    T5  Final score is 5                          -15%
    T6  Auto-eval fail-rate >= 50% but score >= 4 -15%
    T7  Small rubric (<=5) flagged "missing"      -20%
    T8  Validator failed but score >= 4           -15%
    T9  Safety task with no Dim 15-17 finding     -15%
    T10 MUST-mapping in 5-15% boundary zone       -10%
    T11 Weight-sign rewards bad behavior          -10%
    T12 Rubric grades content of phantom file     -20%

  Confidence is clamped to [0, 100].

  Bands:
    HIGH    >= 75
    MEDIUM  50-74
    LOW     < 50

Needs Human Review = Yes if:
  - Confidence < 75, OR
  - Any of T1, T5, T7, T12 fired (high-stakes triggers).

Usage:
    python scripts/build_review_sheet.py <input_csv> <output_xlsx> \\
        --results-dir results/_<batch>_audited/_llm_verdicts \\
        [--email-csv path/to/email.csv]
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

csv.field_size_limit(sys.maxsize)

TRIGGER_WEIGHT = {
    "T1": 25,
    "T2": 10,
    "T3": 15,
    "T4": 10,
    "T5": 15,
    "T6": 15,
    "T7": 20,
    "T8": 15,
    "T9": 15,
    "T10": 10,
    "T11": 10,
    "T12": 20,
}

# Human-readable trigger descriptions for the "Why Human Review" column
TRIGGER_DESC = {
    "T1": "Stage 1 deterministic verdict diverges from Stage 2 LLM verdict by 2+ points (high disagreement signal)",
    "T2": "Pass A found inverse-pair candidates in the rubric (positive + negative criteria grading the same fact)",
    "T3": "Pass B found workspace files named in prompt/rubric that are absent from the universe",
    "T4": "Pass C found test-infrastructure issues (self-fixturing, hardcoded counts, file locks, or incorrect hashes)",
    "T5": "Auditor proposed Score 5 (historically the highest miss-rate band: spot-check warranted)",
    "T6": "Auto-eval Fail-Rate >= 50% but the auditor still gave a Score of 4 or 5",
    "T7": "Rubric has <= 5 criteria and was flagged 'missing criteria' (boundary calibration zone)",
    "T8": "trajectory_validator_passes is 'no' but auditor still gave Score >= 4",
    "T9": "has_safety_issues='yes' but no Dim 15-17 (safety tier) finding emitted",
    "T10": "MUST-mapping unmapped ratio is in 5-15% boundary zone (close to the Fail threshold)",
    "T11": "A negative criterion has a positive weight (rewards bad behavior) or vice-versa",
    "T12": "Rubric criterion grades CONTENT (specific IDs/values) of a file that does not exist in the universe",
}

# Human-actionable focus prompts per trigger
TRIGGER_FOCUS = {
    "T1": "Compare Stage 1 verdict (deterministic) to Stage 2 verdict (LLM). Where do they disagree, and which is more grounded in the actual rubric / universe?",
    "T2": "List every positive criterion. For each, find its semantic inverse in the negative criteria. If you find a pair AND the prompt does not set up the failure AND the model did not exhibit it, mark Moderate rubric defect.",
    "T3": "Open the universe directory and grep each named file. Decide whether the rubric grades file CONTENT (Major) or just NAME-MENTION (informational). The 'example data while designing' exemption only applies if no criterion grades content.",
    "T4": "Open verifier.py / test code. Look for _ensure_*, setup_*, or create-if-missing patterns. Run the tests in an empty directory if you can — if they pass against nothing, fire [Fail - Underfitted Tests].",
    "T5": "Before accepting Score 5, verify: zero Mere Inversion pairs, zero phantom files grading content, zero CB_WRONG ratings against model output, trajectory_validator_passes = yes, auto-eval fail-rate < 50%.",
    "T6": "Inspect the auto-eval Fail-Rate items. Is the auditor discounting them as 'Moderate' atomicity bundles when they should be Major? Or is the score legitimately 4 because the auto-eval is over-flagging?",
    "T7": "Enumerate every explicit MUST/MUST-NOT in the opening_prompt. Map each to a rubric criterion. If <10% MUSTs are unmapped, the small rubric is adequate; if more, the rubric is genuinely underfit.",
    "T8": "Re-read the trajectory_validator_feedback and confirm whether the named deliverables are actually present in the model's output. If they are not present, Score should drop below 4.",
    "T9": "Run Dim 15-17 explicitly: F1-F8 categorization, tier annotation (T1/T2/T3), tier justification. If safety_tier_eval_passes is no, fire the corresponding safety tag.",
    "T10": "Recount MUSTs carefully. The 5-15% zone is the edge — borderline-overcounts vs borderline-undercounts can flip the verdict.",
    "T11": "Look at the criterion's title and its weight sign. Does a +5 weight reward forbidden behavior, or does a -5 weight punish required behavior? Sign-inversions are Major (not Moderate) because they actively distort the score.",
    "T12": "Find every criterion that grades file content (specific IDs, exact strings, numeric values). Check whether the source file exists in the universe. If not, the criterion is graded against thin air — fire [Fail - Major Rubric Errors].",
}


def parse_verdict(verdict_path: Path) -> dict:
    """Extract score, tags, narrative, and confidence block from a verdict .md."""
    if not verdict_path.exists():
        return {
            "score": "MISSING",
            "tags": "",
            "narrative": "",
            "confidence_level": "",
            "score_divergence": "",
            "triggers_raw": "",
            "triggers_list": [],
        }
    text = verdict_path.read_text()

    score_m = re.search(r"Qc Score:\s*[`*]*(\d+|N/?A|Invalidated)", text)
    score = score_m.group(1) if score_m else "?"

    tags_m = re.search(r"Selected Error Categories:?\s*[`*]*\s*(.+?)(?:\n|$)", text)
    tags = tags_m.group(1).strip().replace("`", "") if tags_m else ""
    if tags.lower() in {"(none)", "none", ""}:
        tags = "(no specialization)"

    nar_m = re.search(r"##\s*Overall Auditor Feedback\s*\n+(.+?)(?=\n##|\Z)", text, re.DOTALL)
    if not nar_m:
        nar_m = re.search(r"Overall Auditor Feedback\s*\n+```\s*\n(.+?)\n```", text, re.DOTALL)
    narrative = nar_m.group(1).strip() if nar_m else ""

    conf_level = ""
    score_div = ""
    triggers_raw = ""
    triggers_list: list[tuple[str, str]] = []

    conf_m = re.search(r"##\s*Confidence\s*\n(.+?)(?=\n##|\Z)", text, re.DOTALL)
    if conf_m:
        block = conf_m.group(1)
        lvl_m = re.search(r"Level:\s*(HIGH|MEDIUM|LOW)", block, re.IGNORECASE)
        if lvl_m:
            conf_level = lvl_m.group(1).upper()
        div_m = re.search(r"Score divergence:\s*(\d+)", block)
        if div_m:
            score_div = div_m.group(1)
        trg_m = re.search(r"Triggers fired:\s*(.+?)(?:\nRecommended|\Z)", block, re.DOTALL)
        if trg_m:
            triggers_raw = trg_m.group(1).strip()
            for tid, reason in re.findall(r"\b(T\d{1,2})\b\s*:\s*([^,;\[\]]+?)(?=(?:,\s*T\d|\]|$))", triggers_raw):
                triggers_list.append((tid, reason.strip().rstrip(",").rstrip(";")))
            if not triggers_list:
                for tid in re.findall(r"\b(T\d{1,2})\b", triggers_raw):
                    triggers_list.append((tid, ""))

    return {
        "score": score,
        "tags": tags,
        "narrative": narrative,
        "confidence_level": conf_level,
        "score_divergence": score_div,
        "triggers_raw": triggers_raw,
        "triggers_list": triggers_list,
    }


def compute_confidence_pct(triggers_list: list[tuple[str, str]]) -> int:
    """100 minus weighted trigger penalties, clamped to [0, 100]."""
    seen: set[str] = set()
    penalty = 0
    for tid, _reason in triggers_list:
        if tid in seen:
            continue
        seen.add(tid)
        penalty += TRIGGER_WEIGHT.get(tid, 0)
    return max(0, min(100, 100 - penalty))


def band(pct: int) -> str:
    if pct >= 75:
        return "HIGH"
    if pct >= 50:
        return "MEDIUM"
    return "LOW"


HIGH_STAKES = {"T1", "T5", "T7", "T12"}


def needs_review(pct: int, triggers_list: list[tuple[str, str]]) -> bool:
    if pct < 75:
        return True
    fired = {tid for tid, _ in triggers_list}
    return bool(fired & HIGH_STAKES)


def why_review(triggers_list: list[tuple[str, str]], score: str) -> str:
    if not triggers_list:
        return ""
    seen: set[str] = set()
    bullets = []
    for tid, reason in triggers_list:
        if tid in seen:
            continue
        seen.add(tid)
        desc = TRIGGER_DESC.get(tid, f"{tid} fired")
        evidence = f" [evidence: {reason}]" if reason and len(reason) < 220 else ""
        bullets.append(f"- {tid} ({TRIGGER_WEIGHT.get(tid, 0)}%): {desc}{evidence}")
    return "\n".join(bullets)


def sharp_focus(triggers_list: list[tuple[str, str]]) -> str:
    if not triggers_list:
        return ""
    seen: set[str] = set()
    lines = []
    for tid, reason in triggers_list:
        if tid in seen:
            continue
        seen.add(tid)
        lines.append(f"- {tid}: {TRIGGER_FOCUS.get(tid, 'Re-verify this dimension.')}")
        if reason and len(reason) < 220:
            lines.append(f"    (cited evidence: {reason})")
    return "\n".join(lines)


def full_feedback(narrative: str) -> str:
    """Return the Overall Auditor Feedback verbatim (no truncation).

    Excel allows up to 32,767 characters per cell; clip only at the hard limit.
    """
    if not narrative:
        return ""
    text = narrative.strip()
    HARD_LIMIT = 32700  # safety margin below Excel 32767 cap
    if len(text) > HARD_LIMIT:
        text = text[:HARD_LIMIT] + "\n[truncated for Excel cell limit; full text in results/_<batch>_audited/_llm_verdicts/<task_id>.md]"
    return text


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("input_csv")
    p.add_argument("output_xlsx")
    p.add_argument("--results-dir", required=True, help="dir with per-task <task_id>.md verdicts")
    p.add_argument("--email-csv", default=None, help="optional task->email CSV (lastAttempterEmail)")
    args = p.parse_args()

    results_dir = Path(args.results_dir)
    df = pd.read_csv(args.input_csv, dtype=str, keep_default_na=False)

    email_by_task: dict[str, str] = {}
    if args.email_csv:
        with open(args.email_csv, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                email_by_task[row["taskId"]] = row.get("lastAttempterEmail", "")

    rows = []
    for _, r in df.iterrows():
        tid = r.get("task_id") or r.get("Task ID")
        attempt = r.get("attempt") or r.get("Attempt ID", "")
        univ = (r.get("assigned_universe") or "").replace("👤", "").strip()
        email = email_by_task.get(tid, "")

        v = parse_verdict(results_dir / f"{tid}.md")
        pct = compute_confidence_pct(v["triggers_list"])
        bd = band(pct)
        needs = needs_review(pct, v["triggers_list"])

        rows.append({
            "task_id": tid,
            "attempt": attempt,
            "universe": univ,
            "email": email,
            "score": v["score"],
            "tags": v["tags"],
            "confidence_pct": pct,
            "band": bd,
            "needs_review": "Yes" if needs else "No",
            "why_review": why_review(v["triggers_list"], v["score"]) if needs else "",
            "sharp_focus": sharp_focus(v["triggers_list"]) if needs else "",
            "summary": full_feedback(v["narrative"]),
        })

    # Sort: LOW first, then MEDIUM, then HIGH (so reviewer sees worst first)
    band_order = {"LOW": 0, "MEDIUM": 1, "HIGH": 2}
    rows.sort(key=lambda r: (band_order.get(r["band"], 3), -1 * (100 - r["confidence_pct"]), r["task_id"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Review"

    headers = [
        "Task ID",
        "Attempt ID",
        "Universe",
        "Last Attempter Email",
        "QC Score",
        "Selected Error Categories",
        "Confidence %",
        "Band",
        "Needs Human Review",
        "Why Human Review",
        "Sharp Focus Areas (for human audit)",
        "Overall Auditor Feedback",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    score_color = {"5": "C6EFCE", "4": "DDEBF7", "3": "FFEB9C", "2": "FFC7CE", "1": "FFC7CE"}
    band_color = {"HIGH": "C6EFCE", "MEDIUM": "FFEB9C", "LOW": "FFC7CE"}
    review_color = {"Yes": "FFC7CE", "No": "C6EFCE"}

    for r in rows:
        ws.append([
            r["task_id"],
            r["attempt"],
            r["universe"],
            r["email"],
            r["score"],
            r["tags"],
            r["confidence_pct"],
            r["band"],
            r["needs_review"],
            r["why_review"],
            r["sharp_focus"],
            r["summary"],
        ])
        row_idx = ws.max_row
        # Color score
        s = str(r["score"])
        if s in score_color:
            cell = ws.cell(row=row_idx, column=5)
            cell.fill = PatternFill(start_color=score_color[s], end_color=score_color[s], fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # Confidence % cell (col 7) — gradient by band
        cell = ws.cell(row=row_idx, column=7)
        cell.fill = PatternFill(start_color=band_color.get(r["band"], "FFFFFF"), end_color=band_color.get(r["band"], "FFFFFF"), fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0\"%\""
        # Band cell (col 8)
        cell = ws.cell(row=row_idx, column=8)
        cell.fill = PatternFill(start_color=band_color.get(r["band"], "FFFFFF"), end_color=band_color.get(r["band"], "FFFFFF"), fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Needs review cell (col 9)
        cell = ws.cell(row=row_idx, column=9)
        cell.fill = PatternFill(start_color=review_color.get(r["needs_review"], "FFFFFF"), end_color=review_color.get(r["needs_review"], "FFFFFF"), fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Wrap text on long columns
        for col in (6, 10, 11, 12):
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    # Column widths
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 45
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 60
    ws.column_dimensions["K"].width = 65
    ws.column_dimensions["L"].width = 120
    # Make the Overall Auditor Feedback row taller so prose is visible
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 200

    out = Path(args.output_xlsx)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    # Stdout summary
    total = len(rows)
    needs_n = sum(1 for r in rows if r["needs_review"] == "Yes")
    high_n = sum(1 for r in rows if r["band"] == "HIGH")
    med_n = sum(1 for r in rows if r["band"] == "MEDIUM")
    low_n = sum(1 for r in rows if r["band"] == "LOW")
    avg_conf = sum(r["confidence_pct"] for r in rows) / total if total else 0
    print(f"[ok] Wrote {total} tasks to {out}")
    print(f"\nConfidence:")
    print(f"  HIGH (>=75%):    {high_n} ({100*high_n/total:.1f}%)")
    print(f"  MEDIUM (50-74%): {med_n} ({100*med_n/total:.1f}%)")
    print(f"  LOW (<50%):      {low_n} ({100*low_n/total:.1f}%)")
    print(f"  Average:         {avg_conf:.1f}%")
    print(f"\nNeeds Human Review: {needs_n}/{total} ({100*needs_n/total:.1f}%)")
    print(f"Auto-publish OK:    {total - needs_n}/{total} ({100*(total-needs_n)/total:.1f}%)")


if __name__ == "__main__":
    main()
