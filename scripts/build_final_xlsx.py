#!/usr/bin/env python3
"""Build one consolidated final_audit_results.xlsx across all three batches.

Sheets:
  - Summary       — per-batch counts, score distribution, calibration headline
  - All Audits    — every task with batch, task_id, persona, score, tags
  - Human vs Mine — 10 tasks with human ground truth, score Δ, tag overlap
  - Failure Modes — top failure-mode counts across the whole pool
"""
from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/vishal.kushwaha/Documents/Openclaw_RL copy/openclaw_qc_auditor_v1.1")
OUT_PATH = Path("/Users/vishal.kushwaha/Documents/Openclaw_RL copy/final_audit_results.xlsx")

BATCHES = [
    ("test_v3", ROOT / "test_v3.csv", ROOT / "results/_test_v3_audited/_llm_verdicts"),
    ("L12",     ROOT / "L12.csv",     ROOT / "results/_L12_audited/_llm_verdicts"),
    ("test4",   ROOT / "test4.csv",   ROOT / "results/_test4_audited/_llm_verdicts"),
    ("L11",     ROOT / "L11.csv",     ROOT / "results/_L11_audited/_llm_verdicts"),
    ("L4",      ROOT / "L4.csv",      ROOT / "results/_L4_audited/_llm_verdicts"),
]

SCORE_FILL = {
    5: PatternFill("solid", fgColor="C6EFCE"),   # green
    4: PatternFill("solid", fgColor="E2EFDA"),   # light green
    3: PatternFill("solid", fgColor="FFEB9C"),   # yellow
    2: PatternFill("solid", fgColor="FFC7CE"),   # red
    1: PatternFill("solid", fgColor="F8B7BD"),   # darker red
}
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def parse_verdict(md_path: Path) -> dict:
    text = md_path.read_text()
    score_m = re.search(r"Qc Score:\s*(\d)", text)
    tag_m = re.search(r"Selected Error Categories:\s*(.+)", text)
    return {
        "score": int(score_m.group(1)) if score_m else None,
        "tags": tag_m.group(1).strip() if tag_m else "",
    }


def collect_batch_rows(batch_name: str, csv_path: Path, verdicts_dir: Path) -> list[dict]:
    df = pd.read_csv(csv_path)
    df["persona_slug"] = (df["assigned_universe"].fillna("")
                          .str.replace("👤", "", regex=False)
                          .str.replace("🛎️", "", regex=False)
                          .str.strip())
    rows = []
    for _, r in df.iterrows():
        tid = r["task_id"]
        v_path = verdicts_dir / f"{tid}.md"
        if not v_path.exists():
            continue
        v = parse_verdict(v_path)
        rows.append({
            "batch": batch_name,
            "task_id": tid,
            "persona": r["persona_slug"],
            "scenario_type": r.get("scenario_type", ""),
            "safety_tier": str(r.get("safety_tier_annotation", "") or ""),
            "score": v["score"],
            "tags": v["tags"],
        })
    return rows


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def autofit(ws):
    for col in ws.columns:
        # col is a tuple of cells; take the column letter from the first
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(v), 80))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 8), 90)


def main() -> None:
    # collect
    all_rows: list[dict] = []
    for name, csv_path, verdicts_dir in BATCHES:
        all_rows.extend(collect_batch_rows(name, csv_path, verdicts_dir))

    # Human ground truth aggregated from CSV (8) and (9)
    HUMAN = {
        # test_v3 / CSV (8) overlap
        "6a04c1b084dec168370057b3": (2, "[Fail - 15%+ Moderate Rubric Errors], [Fail - Incorrect Justification]"),
        "6a04c1b084dec1683700582a": (4, "[Non-Fail - Up to 15% Moderate Errors]"),
        "6a0383a392db43e0e0910c1f": (4, "[Non-Fail - Up to 15% Moderate Errors], [Non-Fail - Weak Justification]"),
        "6a04c1b084dec168370057a9": (5, "(no specialization)"),
        # CSV (8) additional
        "6a04c1b084dec1683700580d": (4, "[Non-Fail - Minor Failure Description Issues], [Non-Fail - Failure Categorization]"),
        "6a04c1b084dec168370057c8": (5, "(no specialization)"),
        "6a04c1b084dec168370057df": (3, "[Non-Fail - Underfitted Tests], [Non-Fail - Up to 15% Moderate Errors]"),
        "6a04c1b084dec16837005816": (2, "[Fail - Incorrect Justification], [Non-Fail - 5-20% Minor Errors], [Non-Fail - Underfitted Tests], [Non-Fail - Minor Incorrect Evaluations]"),
        # CSV (9) new
        "6a014bdd0932ab06d5ccbc73": (2, "[Fail - 10%+ Major Rubric Errors], [Fail - 15%+ Moderate Rubric Errors], [Non-Fail - Incorrect Tests], [Non-Fail - Minor Incorrect Evaluations]"),
        "6a04c1b084dec168370057b8": (2, "[Fail - Prompt Feasibility with Tools], [Fail - Incorrect Tests], [Fail - Incorrect Justification], [Non-Fail - Up to 15% Moderate Errors]"),
    }

    wb = Workbook()

    # ---------- Sheet 1: Summary ----------
    ws = wb.active
    ws.title = "Summary"
    ws.append(["OpenClaw QC Auditor — Final Audit Results"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Batches audited", "Tasks", "Source CSV"])
    style_header(ws, 3)
    ws.cell(row=3, column=1).fill = HEADER_FILL
    ws.cell(row=3, column=1).font = HEADER_FONT
    row_i = 4
    for name, csv_path, vdir in BATCHES:
        count = sum(1 for r in all_rows if r["batch"] == name)
        ws.cell(row=row_i, column=1, value=name)
        ws.cell(row=row_i, column=2, value=count)
        ws.cell(row=row_i, column=3, value=csv_path.name)
        row_i += 1
    ws.cell(row=row_i, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row_i, column=2, value=len(all_rows)).font = Font(bold=True)
    row_i += 2

    # Score distribution per batch
    ws.cell(row=row_i, column=1, value="Score Distribution").font = Font(bold=True, size=12)
    row_i += 1
    header = ["Batch", "Score 5", "Score 4", "Score 3", "Score 2", "Score 1", "Total"]
    for c, h in enumerate(header, start=1):
        cell = ws.cell(row=row_i, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row_i += 1
    for name, _, _ in BATCHES + [("ALL", None, None)]:
        batch_rows = all_rows if name == "ALL" else [r for r in all_rows if r["batch"] == name]
        ws.cell(row=row_i, column=1, value=name)
        for c, score in enumerate([5, 4, 3, 2, 1], start=2):
            n = sum(1 for r in batch_rows if r["score"] == score)
            cell = ws.cell(row=row_i, column=c, value=n)
            if n:
                cell.fill = SCORE_FILL[score]
        ws.cell(row=row_i, column=7, value=len(batch_rows))
        if name == "ALL":
            for c in range(1, 8):
                ws.cell(row=row_i, column=c).font = Font(bold=True)
        row_i += 1
    row_i += 1

    # Calibration headline
    ws.cell(row=row_i, column=1, value="Calibration vs Human (10 ground-truth tasks)").font = Font(bold=True, size=12)
    row_i += 1
    headers = ["Metric", "Value", "v1.1 Target"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row_i, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row_i += 1
    mine_lookup = {r["task_id"]: r for r in all_rows}
    deltas = []
    for tid, (h_score, _) in HUMAN.items():
        # prefer L12 verdict where available
        candidates = [r for r in all_rows if r["task_id"] == tid]
        candidate = next((r for r in candidates if r["batch"] == "L12"), candidates[0] if candidates else None)
        if candidate and candidate["score"] is not None:
            deltas.append(candidate["score"] - h_score)
    exact = sum(1 for d in deltas if d == 0)
    within1 = sum(1 for d in deltas if abs(d) <= 1)
    off2 = sum(1 for d in deltas if abs(d) >= 2)
    for label, val, tgt in [
        ("Exact score match", f"{exact}/{len(deltas)} ({100*exact/len(deltas):.0f}%)", "93%"),
        ("Within 1 score point", f"{within1}/{len(deltas)} ({100*within1/len(deltas):.0f}%)", "100%"),
        ("Off by 2+", f"{off2}/{len(deltas)} ({100*off2/len(deltas):.0f}%)", "0%"),
    ]:
        ws.cell(row=row_i, column=1, value=label)
        ws.cell(row=row_i, column=2, value=val)
        ws.cell(row=row_i, column=3, value=tgt)
        row_i += 1
    autofit(ws)

    # ---------- Sheet 2: All Audits ----------
    ws = wb.create_sheet("All Audits")
    cols = ["Batch", "Task ID", "Persona", "Scenario Type", "Safety Tier", "QC Score", "Selected Error Categories"]
    ws.append(cols)
    style_header(ws, len(cols))
    # Sort: batch first (insertion order from BATCHES), then score asc, then task_id
    batch_order = {b[0]: i for i, b in enumerate(BATCHES)}
    sorted_rows = sorted(all_rows, key=lambda r: (batch_order[r["batch"]], r["score"] or 9, r["task_id"]))
    for r in sorted_rows:
        ws.append([r["batch"], r["task_id"], r["persona"], r["scenario_type"], r["safety_tier"], r["score"], r["tags"]])
        if r["score"] in SCORE_FILL:
            ws.cell(row=ws.max_row, column=6).fill = SCORE_FILL[r["score"]]
    ws.freeze_panes = "A2"
    autofit(ws)

    # ---------- Sheet 3: Human vs Mine ----------
    ws = wb.create_sheet("Human vs Mine")
    cols = ["Task ID", "Persona", "Batch", "My Score", "Human Score", "Δ (mine - human)", "My Tags", "Human Tags", "Match Type"]
    ws.append(cols)
    style_header(ws, len(cols))
    for tid, (h_score, h_tags) in HUMAN.items():
        candidates = [r for r in all_rows if r["task_id"] == tid]
        candidate = next((r for r in candidates if r["batch"] == "L12"), candidates[0] if candidates else None)
        if not candidate:
            continue
        my_score = candidate["score"]
        delta = (my_score - h_score) if my_score is not None else None
        if delta == 0:
            match = "Exact ✓"
        elif delta is not None and abs(delta) == 1:
            match = "Within 1"
        elif delta is not None and abs(delta) >= 2:
            match = f"Off by {abs(delta)}"
        else:
            match = "?"
        ws.append([tid, candidate["persona"], candidate["batch"], my_score, h_score, delta, candidate["tags"], h_tags, match])
        if my_score in SCORE_FILL:
            ws.cell(row=ws.max_row, column=4).fill = SCORE_FILL[my_score]
        if h_score in SCORE_FILL:
            ws.cell(row=ws.max_row, column=5).fill = SCORE_FILL[h_score]
        if delta == 0:
            ws.cell(row=ws.max_row, column=9).fill = PatternFill("solid", fgColor="C6EFCE")
        elif delta is not None and abs(delta) == 1:
            ws.cell(row=ws.max_row, column=9).fill = PatternFill("solid", fgColor="FFEB9C")
        else:
            ws.cell(row=ws.max_row, column=9).fill = PatternFill("solid", fgColor="FFC7CE")
    ws.freeze_panes = "A2"
    autofit(ws)

    # ---------- Sheet 4: Failure Modes ----------
    ws = wb.create_sheet("Failure Modes")
    cols = ["Tag", "Count", "% of Tasks"]
    ws.append(cols)
    style_header(ws, len(cols))
    tag_counter: Counter[str] = Counter()
    for r in all_rows:
        for tag in re.findall(r"\[[^\[\]]+\]\s*\[[^\[\]]+\]\s*\[([^\[\]]+)\]", r["tags"] or ""):
            tag_counter[f"[{tag}]"] += 1
    total = len(all_rows)
    for tag, n in tag_counter.most_common():
        ws.append([tag, n, f"{100*n/total:.1f}%"])
        if "Fail -" in tag:
            ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor="FFC7CE")
        else:
            ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor="FFEB9C")
    ws.freeze_panes = "A2"
    autofit(ws)

    wb.save(OUT_PATH)
    print(f"[ok] wrote {OUT_PATH}")
    print(f"  batches: {[b[0] for b in BATCHES]}")
    print(f"  total tasks: {len(all_rows)}")
    print(f"  human-validated: {len(HUMAN)}")
    print(f"  exact / within-1 / off-2+: {exact}/{within1}/{off2} of {len(deltas)}")


if __name__ == "__main__":
    main()
