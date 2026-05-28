#!/usr/bin/env python3
"""Apply workspace-grounded recheck output to an audit xlsx.

Reads per-universe recheck JSON files (each: {task_id: {feedback, quality, unnecessary_criteria}})
and updates the xlsx so that:
- The Human-Friendly Feedback column is overwritten with the rechecked feedback.
- A new Task Quality column is added (re-do | quick fix | okay | good), color-coded.

Usage:
    python scripts/apply_recheck.py \
        <input.xlsx> <output.xlsx> \
        results/<batch>_audited/_recheck/*.json
"""

from __future__ import annotations

import argparse
import glob
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


QUALITY_COLOR = {
    "re-do":     "FFC7CE",  # red
    "quick fix": "FFEB9C",  # yellow
    "okay":      "DDEBF7",  # light blue
    "good":      "C6EFCE",  # green
}


def load_recheck(json_paths: list[str]) -> dict[str, dict]:
    merged: dict[str, dict] = {}
    for pattern in json_paths:
        paths = sorted(glob.glob(pattern)) or [pattern]
        for p in paths:
            pp = Path(p)
            if not pp.exists() or pp.name == "_universe_groups.json":
                continue
            with pp.open() as f:
                data = json.load(f)
            print(f"[ok]   loaded {pp.name}: {len(data)} entries")
            merged.update(data)
    return merged


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("input_xlsx")
    p.add_argument("output_xlsx")
    p.add_argument("recheck_jsons", nargs="+",
                   help="One or more JSON files (or glob patterns) with recheck output")
    p.add_argument("--task-id-col", default="Task ID")
    p.add_argument("--feedback-col", default="Human-Friendly Feedback")
    p.add_argument("--quality-col", default="Task Quality")
    args = p.parse_args()

    recheck = load_recheck(args.recheck_jsons)
    if not recheck:
        sys.exit("[err] no recheck data loaded")
    print(f"[ok]   total: {len(recheck)} task entries")

    wb = load_workbook(args.input_xlsx)
    ws = wb["Audit Results"] if "Audit Results" in wb.sheetnames else wb.active

    headers = [c.value for c in ws[1]]
    try:
        tid_idx = headers.index(args.task_id_col) + 1
    except ValueError:
        sys.exit(f"[err] task ID column not found: {args.task_id_col!r}")

    feedback_idx = headers.index(args.feedback_col) + 1 if args.feedback_col in headers else None
    quality_idx = headers.index(args.quality_col) + 1 if args.quality_col in headers else None

    # Add columns if missing
    if feedback_idx is None:
        feedback_idx = ws.max_column + 1
        ws.cell(row=1, column=feedback_idx, value=args.feedback_col)
    if quality_idx is None:
        quality_idx = ws.max_column + 1
        ws.cell(row=1, column=quality_idx, value=args.quality_col)

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for col_idx in (feedback_idx, quality_idx):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    matched = 0
    missing: list[str] = []
    quality_count = {"re-do": 0, "quick fix": 0, "okay": 0, "good": 0}
    for r in range(2, ws.max_row + 1):
        tid = str(ws.cell(row=r, column=tid_idx).value or "").strip()
        entry = recheck.get(tid)
        if entry is None:
            missing.append(tid)
            continue
        fb = ws.cell(row=r, column=feedback_idx, value=entry.get("feedback", ""))
        fb.alignment = Alignment(wrap_text=True, vertical="top")

        q = entry.get("quality", "")
        qcell = ws.cell(row=r, column=quality_idx, value=q)
        qcell.alignment = Alignment(horizontal="center", vertical="center")
        qcell.font = Font(bold=True)
        if q in QUALITY_COLOR:
            qcell.fill = PatternFill(start_color=QUALITY_COLOR[q],
                                     end_color=QUALITY_COLOR[q], fill_type="solid")
            quality_count[q] += 1

        matched += 1

    # Column widths
    ws.column_dimensions[ws.cell(row=1, column=feedback_idx).column_letter].width = 90
    ws.column_dimensions[ws.cell(row=1, column=quality_idx).column_letter].width = 14

    # Refresh Summary sheet with the new quality counts
    if "Summary" in wb.sheetnames:
        ws2 = wb["Summary"]
        # Append quality distribution at end
        ws2.append([])
        ws2.append(["Task Quality"])
        ws2.append(["Quality", "Count", "% of total"])
        for label in ("re-do", "quick fix", "okay", "good"):
            n = quality_count[label]
            pct = f"{100*n/matched:.1f}%" if matched else "0%"
            ws2.append([label, n, pct])

    out = Path(args.output_xlsx)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print(f"\n[done] Wrote {out} with {matched}/{ws.max_row - 1} updates.")
    print("Quality distribution:")
    for label in ("re-do", "quick fix", "okay", "good"):
        n = quality_count[label]
        pct = 100 * n / matched if matched else 0
        print(f"  {label:>10}: {n:>3} ({pct:.0f}%)")
    if missing:
        print(f"\n[warn] {len(missing)} tasks have no recheck entry. Sample: {missing[:5]}")


if __name__ == "__main__":
    main()
